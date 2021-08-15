import sys
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdftypes import resolve1
import openpyxl
from tkinter import filedialog
from tkinter import messagebox
import os

messagebox.showinfo("Note", "Please choose the folder where the PDF files are")
PDFpath = filedialog.askdirectory()

workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "data"
column = 2
worksheet.cell(1, 1, "DATA ENTRY")
ExcelItems = ['ENTRY NUMBER', 'Entry Date', 'Branch', 'Audit Date', 'Entry Writer', 'Entry Summary CBP Form Included in File', 'US Goods Returned Affidavit for >$10K', 'Commercial Invoice Included in File', 'FTA Supporting Documents', 'Rates invoice in File', 'Importer Declaration (TSCA, Footwear, etc.)', 'Packing List in File', 'Bill of Lading/Airway Bill in File', 'ADD/CVD Declaration Disclaimer in File', 'HTS Codes Correct as Listed on Invoice', 'Deductions/Additions taken based on INCOTERMS', 'C/O entered matches commercial invoice', 'Correct currency along with rate of exchange', 'Assists Declared Properly', 'SPI indicator entered as Applicable', 'Buying/Selling Comissions Properly Declared', 'Correct Invoice Values Indicated', 'If Product Coding, Correct PN Used', 'Description Sufficient to Classify', 'CO Indicated on Invoice', 'Quantity indicated', 'Unit Value and Total Value Indicated', 'Currency of Transaction Indicated', 'INCOTERM Indicated', 'Importer/Consignee Match on Invoice', 'All HTS Codes from Invoice included in 7501']
for x in range(len(ExcelItems)):
    worksheet.cell(x+2,1,ExcelItems[x])

for root, dirs, files in os.walk(PDFpath):
    for fn in files:
        row = 2
        filename = sys.argv[0]
        fp = open(root + '/' + fn, 'rb')

        worksheet.cell(1, column, "RESPONSE - " + fn[:-4])
        parser = PDFParser(fp)
        doc = PDFDocument(parser)
        fields = resolve1(doc.catalog['AcroForm'])['Fields']

        for i in fields:
            field = resolve1(i)
            name, value = field.get('T'), field.get('V')
            worksheet.cell(row,column,value)
            row += 1

        column += 1

messagebox.showinfo("Note", "Please choose the folder where put the Excels")
Excelpath = filedialog.askdirectory()
workbook.save(filename= Excelpath + '/Result.xlsx')







